from flask import Flask, request, jsonify
import openpyxl
from flask_cors import CORS
import os


app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Load Excel file and extract user data
def load_users_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        username, password, key, full_name, team = row[:5]

        # check if some values are empty and if so, stop the loop
        if username is None:
            break

        users.append({"username": username, "password": password, "key": key, "full_name": full_name, "team": team})
    return users


def append_to_excel(file_path, data):
    try:
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Remove empty rows
        rows_to_keep = []
        for row in sheet.iter_rows():
            if any(cell.value is not None for cell in row):  # Check if the row has any non-empty cell
                rows_to_keep.append([cell.value for cell in row])  # Keep the row data

        # Clear the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None

        # Write back the non-empty rows
        for i, row_data in enumerate(rows_to_keep, start=1):
            for j, value in enumerate(row_data, start=1):
                sheet.cell(row=i, column=j).value = value

        # Append the new data as a new row
        sheet.append(data)

        # Save the workbook
        workbook.save(file_path)
        print("Data appended successfully.")
    except Exception as e:
        print(f"Error while writing to Excel: {e}")
        raise

def get_file_path(file_name):
    current_file_path = os.path.abspath(__file__)
    api_folder_path = os.path.dirname(current_file_path)
    project_folder_path = os.path.dirname(api_folder_path)

    # return os.path.join(project_folder_path, file_name)

    return file_name

# Load users from the Excel file
USERS = load_users_from_excel(get_file_path("login_data.xlsx"))


def get_first_name(key):
    for user in USERS:
        if user["key"] == key:
            return user["username"]


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username")
    password = int(data.get("password"))

    # Find user by username and password
    user = next((u for u in USERS if u["username"] == username and u["password"] == password), None)

    if user:
        return jsonify({"success": True, "key": user["key"],  "full_name": user["full_name"], "team": user["team"]}), 200
    else:
        return jsonify({"success": False, "message": "Invalid credentials"}), 401

@app.route("/api/requestmedic", methods=["POST"])
def requestmedic():
    data = request.get_json()

    key = data.get("key")
    full_name = data.get("full_name")
    request_type = data.get("request_type")
    request_date = data.get("request_date")
    request_message = data.get("request_message")

    append_to_excel(get_file_path("medic_requests.xlsx"), [key, full_name, request_type, request_date, request_message])

    return jsonify({"success": True, "key": key}), 200

@app.route("/api/requestlogistic", methods=["POST"])
def requestlogistic():
    data = request.get_json()

    key = data.get("key")
    full_name = data.get("full_name")
    request_message = data.get("request_message")

    append_to_excel(get_file_path("logistic_requests.xlsx"), [key, full_name, request_message])

    return jsonify({"success": True, "key": key}), 200

@app.route("/api/missionupdate", methods=["POST"])
def missionupdate():
    data = request.get_json()

    key = int(data.get("key"))
    full_name = data.get("full_name")
    first_name = get_first_name(key)
    team = data.get("team")
    mission_id = data.get("mission_id")
    is_done = data.get("is_done")

    # open the excel file and change the value of the appropriate cell
    wb = openpyxl.load_workbook(get_file_path(f"team{team}.xlsx"))
    sheet = wb.active

    # find the appropriate column according to the name
    first_row = sheet[1]
    column_number = 0
    for column_index, cell in enumerate(first_row):
        if cell.value == first_name:
            column_number = column_index
            break

    # change the value of the appropriate cell
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row[0] == mission_id:
            sheet.cell(i + 2, column_number + 1).value = is_done

    # save the changes
    wb.save(get_file_path(f"team{team}.xlsx"))

    return jsonify({"success": True, "key": key, "full_name": full_name}), 200


@app.route("/api/getmissions", methods=["POST"])
def getmissions():
    data = request.get_json()

    key = int(data.get("key"))
    team = data.get("team")
    first_name = get_first_name(key)

    # open the excel file and find the appropriate column
    wb = openpyxl.load_workbook(get_file_path(f"team{team}.xlsx"))
    sheet = wb.active

    first_row = sheet[1]
    column_number = 0
    for column_index, cell in enumerate(first_row):
        if cell.value == first_name:
            column_number = column_index
            break

    # get the missions
    missions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        mission_id, mission_name, mission_date = row[:3]
        is_done = row[column_number]
        missions.append({"mission_id": mission_id, "mission_name": mission_name, "mission_date": mission_date, "is_done": is_done})

    return jsonify({"success": True, "key": key, "missions": missions}), 200


@app.route("/")
def home():
    return "Server is running!"


if __name__ == "__main__":
    app.run(debug=False, port=5000)
